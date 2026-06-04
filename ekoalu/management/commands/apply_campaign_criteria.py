"""Applique les criteres affines de Richard (Excel CAMPAGNES_CRITERES.xlsx)
dans Campaign.campaign_objective, que lit le juge Claude a la qualification.

Lit les colonnes JAUNES de l'Excel (ACTION / CRITERES AFFINES / GEO / MOTS-CLES)
et injecte un bloc idempotent "## Ciblage affine (Richard)" par campagne.

- CRITERES AFFINES + GEO -> bloc dans campaign_objective (applique).
- ACTION 'Supprimer'/'Pause' -> seulement SIGNALE (jamais destructif en auto).
- MOTS-CLES a ajouter -> seulement signale (la generation reste automatique).

Usage :
    python manage.py apply_campaign_criteria --dry-run
    python manage.py apply_campaign_criteria
    python manage.py apply_campaign_criteria --source "C:\\...\\CAMPAGNES_CRITERES.xlsx"

Idempotent : re-passe -> remplace le bloc existant, n'empile pas.
"""
from __future__ import annotations

import os

from django.core.management.base import BaseCommand, CommandError

DEFAULT_SOURCE = r"C:\Users\RI.GROS\Documents\CLAUDE\prospection-ia\CAMPAGNES_CRITERES.xlsx"

# Colonnes (1-indexed) de l'onglet "Campagnes".
COL_NAME, COL_ACTION, COL_CRIT, COL_GEO, COL_KW = 1, 9, 10, 11, 12


class Command(BaseCommand):
    help = "Injecte les criteres affines de l'Excel dans Campaign.campaign_objective."

    def add_arguments(self, parser):
        parser.add_argument("--source", default=DEFAULT_SOURCE, help="Chemin de l'Excel")
        parser.add_argument("--sheet", default="Campagnes", help="Nom de l'onglet")
        parser.add_argument("--dry-run", action="store_true", help="Affiche sans ecrire")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError("openpyxl requis : pip install openpyxl") from e

        source = opts["source"]
        if not os.path.exists(source):
            raise CommandError(f"Excel introuvable : {source}")

        from linkedin.models import Campaign
        from ekoalu.lead_routing.criteria import build_refined_objective, normalize_geo

        wb = openpyxl.load_workbook(source, data_only=True)
        ws = wb[opts["sheet"]]

        # Lookup campagne par nom (avec et sans prefixe "EKOALU - ").
        by_name = {}
        for c in Campaign.objects.all():
            by_name[c.name] = c
            by_name[c.name.replace("EKOALU - ", "")] = c

        dry = opts["dry_run"]
        applied = unmatched = skipped = 0
        notes: list[str] = []

        for row in range(2, ws.max_row + 1):
            raw_name = (ws.cell(row, COL_NAME).value or "").strip()
            if not raw_name:
                continue
            campaign = by_name.get(raw_name) or by_name.get(f"EKOALU - {raw_name}")
            if campaign is None:
                unmatched += 1
                notes.append(f"  [?] Campagne non trouvee : '{raw_name}'")
                continue

            action = (ws.cell(row, COL_ACTION).value or "").strip()
            crit = (ws.cell(row, COL_CRIT).value or "").strip()
            geo = (ws.cell(row, COL_GEO).value or "").strip()
            kw = (ws.cell(row, COL_KW).value or "").strip()

            if action and any(w in action.lower() for w in ("supprim", "pause")):
                notes.append(f"  [!] '{campaign.name}' ACTION='{action}' -> a traiter a la main (non destructif auto)")
            if kw:
                notes.append(f"  [+] '{campaign.name}' mots-cles suggeres : {kw}")

            if not crit and not normalize_geo(geo):
                skipped += 1
                continue

            new_obj = build_refined_objective(campaign.campaign_objective, crit, geo)
            if new_obj == campaign.campaign_objective:
                skipped += 1
                continue

            applied += 1
            if not dry:
                campaign.campaign_objective = new_obj
                campaign.save(update_fields=["campaign_objective"])

        verb = "seraient mis a jour" if dry else "mis a jour"
        self.stdout.write(f"Campagnes : {applied} {verb}, {skipped} inchangees, {unmatched} non trouvees.")
        for n in notes:
            self.stdout.write(n)
        if dry:
            self.stdout.write("(dry-run : rien ecrit)")
